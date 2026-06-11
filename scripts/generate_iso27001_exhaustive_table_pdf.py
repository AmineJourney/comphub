from collections import Counter, defaultdict
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "backend" / "library" / "data" / "iso27001-2022-app.yaml"
OUTPUT = ROOT / "ISO27001_CHECKLIST_TABLE_EXHAUSTIVE_FR.xlsx"


def load_controls():
    data = yaml.safe_load(SOURCE.read_text(encoding="utf-8"))
    return data["controls"]


def section_key(code: str):
    code = str(code)
    if code.startswith("A."):
        parts = code.split(".")
        return f"Annexe {parts[0]}.{parts[1]}"
    return f"Clause {code.split('.')[0]}"


def sort_key(code: str):
    code = str(code)
    if code.startswith("A."):
        _, a, b = code.split(".")
        return (1, int(a), int(b))
    return (0, tuple(int(part) for part in code.split(".")))


def get_fr_text(value):
    if isinstance(value, dict):
        return value.get("fr") or value.get("en") or ""
    return str(value or "")


def management_evidence(code: str) -> str:
    section = code.split(".")[0]
    mapping = {
        "4": "Analyse de contexte, registre des parties interessees, document de perimetre, cartographie des processus, validations de direction.",
        "5": "Politique approuvee, comptes-rendus de comite, RACI, fiches de poste, lettres de mission, preuves d'engagement de la direction.",
        "6": "Methodologie de risque, registre des risques, plan de traitement, SoA, objectifs securite, indicateurs, validations formelles.",
        "7": "Plan de formation, preuves de sensibilisation, plan de communication, procedure documentaire, registre documentaire, preuves de competence.",
        "8": "Procedures operationnelles, tickets, comptes-rendus de revue, journaux, preuves d'execution, suivi des actions de traitement.",
        "9": "Tableaux de bord, KPI, programme d'audit, rapports d'audit, registres de non-conformites, revue de direction.",
        "10": "Registre des non-conformites, analyses de cause, actions correctives, suivi d'efficacite, plan d'amelioration continue.",
    }
    return mapping.get(section, "Documents, enregistrements, validations et preuves d'application correspondants.")


def annex_evidence(code: str) -> str:
    prefix = ".".join(code.split(".")[:2])
    mapping = {
        "A.5": "Politiques, procedures, registres, clauses contractuelles, comptes-rendus de revue, tickets d'incident, preuves d'application organisationnelle.",
        "A.6": "Dossiers RH, contrats, NDA, supports de sensibilisation, feuilles de presence, preuves d'onboarding et offboarding.",
        "A.7": "Journaux d'acces physique, registres visiteurs, CCTV, plans de site, maintenance, certificats de destruction, preuves de securite physique.",
        "A.8": "Configurations techniques, journaux IAM, rapports EDR/SIEM, rapports de sauvegarde, scans de vulnerabilites, tickets de changement, preuves de surveillance.",
    }
    return mapping.get(prefix, "Procedures, journaux, configurations et preuves operationnelles correspondantes.")


def verification_questions(code: str, title: str, description: str) -> str:
    if code.startswith("A."):
        return (
            f"L'organisation a-t-elle mis en oeuvre le controle '{title}' sur le perimetre du SMSI ?\n"
            f"Le controle est-il formalise, applique, revu periodiquement et soutenu par des preuves ?\n"
            f"Point d'attention : {description}"
        )
    return (
        f"L'exigence '{title}' est-elle formalisee, comprise et appliquee dans le fonctionnement du SMSI ?\n"
        f"Des preuves montrent-elles sa mise en oeuvre effective, son suivi et sa revue ?\n"
        f"Point d'attention : {description}"
    )


def section_intro(section: str) -> str:
    intros = {
        "Clause 4": "Exigences de cadrage du SMSI : contexte, parties interessees, perimetre et systeme de management.",
        "Clause 5": "Exigences de leadership : engagement de la direction, politique et responsabilites.",
        "Clause 6": "Exigences de planification : risques, opportunites, traitement et objectifs securite.",
        "Clause 7": "Exigences de support : ressources, competence, sensibilisation, communication et documentation.",
        "Clause 8": "Exigences de fonctionnement : execution des processus, evaluation et traitement des risques en exploitation.",
        "Clause 9": "Exigences d'evaluation des performances : mesure, audit interne et revue de direction.",
        "Clause 10": "Exigences d'amelioration : non-conformites, actions correctives et amelioration continue.",
        "Annexe A.5": "Controles organisationnels de l'Annexe A.",
        "Annexe A.6": "Controles humains de l'Annexe A.",
        "Annexe A.7": "Controles physiques de l'Annexe A.",
        "Annexe A.8": "Controles technologiques de l'Annexe A.",
    }
    return intros.get(section, "")


def build_rows(controls):
    grouped = defaultdict(list)
    for control in controls:
        code = str(control["code"])
        title = get_fr_text(control.get("title"))
        description = get_fr_text(control.get("description")) or get_fr_text(control.get("guidance"))
        evidence = annex_evidence(code) if code.startswith("A.") else management_evidence(code)
        grouped[section_key(code)].append(
            [
                code,
                title,
                description,
                verification_questions(code, title, description),
                evidence,
                "",
                "",
            ]
        )

    for section in grouped:
        grouped[section].sort(key=lambda row: sort_key(row[0]))
    return grouped


def add_overview_sheet(workbook, controls, ordered_sections):
    ws = workbook.active
    ws.title = "Vue d'ensemble"
    ws["A1"] = "ISO/IEC 27001 - Checklist Tabulaire Exhaustive en Francais"
    ws["A1"].font = Font(size=14, bold=True, color="16324A")
    ws["A3"] = (
        "Ce classeur est genere a partir du fichier ISO 27001 present dans l'application. "
        "Il couvre l'ensemble des clauses 4 a 10 ainsi que les 93 controles de l'Annexe A."
    )
    ws["A5"] = "Nombre total d'elements"
    ws["B5"] = len(controls)

    counts = Counter(section_key(str(control["code"])) for control in controls)
    ws["A7"] = "Section"
    ws["B7"] = "Nombre"
    ws["A7"].font = Font(bold=True, color="FFFFFF")
    ws["B7"].font = Font(bold=True, color="FFFFFF")
    ws["A7"].fill = PatternFill("solid", fgColor="3E5F8A")
    ws["B7"].fill = PatternFill("solid", fgColor="3E5F8A")

    row = 8
    for section in ordered_sections:
        if section in counts:
            ws.cell(row=row, column=1, value=section)
            ws.cell(row=row, column=2, value=counts[section])
            row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12


def format_sheet(ws, table_start_row, table_end_row):
    widths = {
        "A": 14,
        "B": 38,
        "C": 46,
        "D": 60,
        "E": 48,
        "F": 16,
        "G": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=table_start_row, max_row=table_end_row, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for cell in ws[table_start_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3E5F8A")

    ws.freeze_panes = f"A{table_start_row + 1}"

    table_ref = f"A{table_start_row}:G{table_end_row}"
    table_name = f"Checklist_{ws.title.replace(' ', '_').replace('.', '_')}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def add_section_sheet(workbook, section, rows):
    ws = workbook.create_sheet(title=section[:31])
    ws["A1"] = section
    ws["A1"].font = Font(size=13, bold=True, color="16324A")
    ws["A2"] = section_intro(section)
    ws["A2"].alignment = Alignment(wrap_text=True)

    headers = [
        "Reference",
        "Exigence / controle",
        "Description / objectif",
        "Questions de verification",
        "Preuves attendues",
        "Statut",
        "Commentaires",
    ]
    start_row = 4
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=idx, value=header)

    row_idx = start_row + 1
    for row in rows:
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        row_idx += 1

    format_sheet(ws, start_row, row_idx - 1)


def main():
    controls = load_controls()
    grouped_rows = build_rows(controls)

    ordered_sections = [
        "Clause 4",
        "Clause 5",
        "Clause 6",
        "Clause 7",
        "Clause 8",
        "Clause 9",
        "Clause 10",
        "Annexe A.5",
        "Annexe A.6",
        "Annexe A.7",
        "Annexe A.8",
    ]

    workbook = Workbook()
    add_overview_sheet(workbook, controls, ordered_sections)

    for section in ordered_sections:
        rows = grouped_rows.get(section, [])
        if rows:
            add_section_sheet(workbook, section, rows)

    workbook.save(OUTPUT)
    print(f"Generated {OUTPUT}")


if __name__ == "__main__":
    main()
